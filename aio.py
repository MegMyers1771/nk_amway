from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Alignment, Side
from openpyxl.cell.cell import Cell
from datetime import datetime

def copy_row_with_merges(ws, source_row, target_row, col_start=2, col_end=40):
    for col in range(col_start, col_end + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)

        if not isinstance(source_cell, Cell):
            continue

        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell._style = source_cell._style

    for rng in list(ws.merged_cells.ranges):
        if rng.min_row == source_row and rng.max_row == source_row:
            ws.merge_cells(
                f"{get_column_letter(rng.min_col)}{target_row}:{get_column_letter(rng.max_col)}{target_row}"
            )

def insert_items(ws, items, start_row=14):
    ad_value = ws["AD14"].value
    for idx, item in enumerate(items):
        row = start_row + idx
        ws[f"B{row}"] = idx + 1
        ws[f"D{row}"] = item["article"]
        ws[f"H{row}"] = item["name"]
        ws[f"AB{row}"] = item["quantity"]
        ws[f"AG{row}"] = f"{item['price']},00"
        ws[f"AK{row}"] = item["quantity"] * item["price"]
        ws[f"AD{row}"] = ad_value

def fill_header(ws, invoice_number, date, manager, buyer, address, payment_method):
    formatted_date = date
    ws["B2"] = f"Расходная накладная № {invoice_number} от {formatted_date}"
    ws["G4"] = manager
    ws["G6"] = buyer
    ws["G8"] = address
    ws["G10"] = payment_method

def insert_summary_block(ws, items, start_row):
    total = sum(item["quantity"] * item["price"] for item in items)
    count = len(items)

    s = start_row
    q = s + 2
    o = s + 6
    k = s + 7

    bold_font = Font(name='Arial', size=10, bold=True)
    left_align = Alignment(horizontal='left')
    center_align = Alignment(horizontal='center')

    # AGs–AJs: "Итого:"
    ws.merge_cells(f"AG{s}:AJ{s}")
    cell_itogo = ws[f"AG{s}"]
    cell_itogo.value = "Итого:"
    cell_itogo.font = bold_font
    cell_itogo.alignment = center_align

    # AKs–ANs: сумма (в формате с запятой и двумя нулями)
    ws.merge_cells(f"AK{s}:AN{s}")
    cell_sum = ws[f"AK{s}"]
    cell_sum.value = f"{total:.2f}".replace(".", ",")
    cell_sum.font = bold_font
    cell_sum.alignment = center_align

    # Bq–ANq: описание строки
    ws.merge_cells(f"B{q}:AN{q}")
    desc_cell = ws[f"B{q}"]
    desc_cell.value = f"Всего наименований {count}, на сумму {total:.2f}".replace(".", ",") + " руб."
    desc_cell.alignment = left_align

    # Подписи
    ws.merge_cells(f"B{o}:E{o}")
    otpus_cell = ws[f"B{o}"]
    otpus_cell.value = "Отпустил"
    otpus_cell.font = bold_font
    otpus_cell.alignment = center_align

    ws.merge_cells(f"U{o}:X{o}")
    poluch_cell = ws[f"U{o}"]
    poluch_cell.value = "Получил"
    poluch_cell.font = bold_font
    poluch_cell.alignment = center_align

    ws.merge_cells(f"H{o}:S{o}")
    podpis1 = ws[f"H{o}"]
    podpis1.value = "____________________"
    podpis1.alignment = center_align

    ws.merge_cells(f"Z{o}:AK{o}")
    podpis2 = ws[f"Z{o}"]
    podpis2.value = "____________________"
    podpis2.alignment = center_align


def process_invoice_template(
        template_path,
        output_path,
        items,
        invoice_number,
        date,
        manager,
        buyer,
        address,
        payment_method
):
    wb = load_workbook(template_path)
    ws = wb.active

    num_items = len(items)

    # 1. Копируем строки таблицы
    for i in range(num_items):
        copy_row_with_merges(ws, 16, 17 + i)

    for i in range(num_items - 2):
        copy_row_with_merges(ws, 15, 15 + i + 1)

    # 4. Вставляем данные
    insert_items(ws, items)

    # 5. Вставляем сумму и описание + подписи
    insert_summary_block(ws, items, start_row=14 + len(items) + 2)

    # 6. Заполняем шапку
    fill_header(ws, invoice_number, date, manager, buyer, address, payment_method)

    # 8. Сохраняем результат
    wb.save(output_path)
